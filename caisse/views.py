import django.db.models.deletion
from django.http import HttpResponse, JsonResponse
from django.core.handlers.wsgi import WSGIRequest
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.serializers import serialize
from django.db import models  # Ajoutez cette ligne
import json
from decimal import Decimal
from .models import Categorie, Personnel, Fournisseur, OperationEntrer, OperationSortir, Beneficiaire
from .forms import FournisseurForm, PersonnelForm, CategorieForm, OperationEntrerForm, OperationSortirForm
from django.db.models import Sum, Count
from django.core.paginator import Paginator
from django.db.models.functions import TruncYear, TruncMonth
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, date
from datetime import timedelta
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import get_user_model
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import update_session_auth_hash
from django.template import loader
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.urls import reverse
from itertools import chain
from operator import attrgetter
from django.core.paginator import Paginator
from .models import UserActivity
from functools import wraps
from babel.dates import format_date
from django.db.models import F
from collections import defaultdict
from django.utils.dateparse import parse_date

import urllib.parse


User = get_user_model()

def is_admin(user):
    return user.is_staff

def superuser_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_superuser:
            return redirect('caisse:index')  # Remplacez 'index' par le nom de votre URL d'index
        return view_func(request, *args, **kwargs)
    return _wrapped_view

# Vues principales
@login_required(login_url="accounts:login_user")
def index(request: WSGIRequest):
    """Vue du tableau de bord"""
    today = timezone.now()
    
    # Obtenir l'année sélectionnée (par défaut, l'année en cours)
    selected_year = int(request.GET.get('year', today.year))  
    
    depense_month = "12" if selected_year != today.year else f"{today.month:02d}"

    # Calculer le premier et dernier jour du mois actuel
    # today = date.today()
    first_day_of_month = datetime(today.year, today.month, 1)
    last_day_of_month = datetime(today.year, today.month + 1, 1) - timedelta(days=1)  # Dernier jour du mois
        
    # Calculer les totaux du mois actuel
    total_entrees_mois = OperationEntrer.objects.filter(
        date_transaction__gte=first_day_of_month,
        date_transaction__lte=last_day_of_month
    ).aggregate(Sum('montant'))['montant__sum'] or Decimal('0')
    
    total_sorties_mois = OperationSortir.objects.filter(
        date_de_sortie__gte=first_day_of_month,
        date_de_sortie__lte=last_day_of_month
    ).aggregate(Sum('montant'))['montant__sum'] or Decimal('0')
    
    
    # Calculer le premier et dernier jour de l'année sélectionnée
    first_day_of_year = datetime(selected_year, 1, 1)
    last_day_of_year = datetime(selected_year, 12, 31)
    
    # Calculer les totaux de l'année sélectionnée
    # total_entrees_mois = OperationEntrer.objects.filter(
    #     date_transaction__gte=first_day_of_year,
    #     date_transaction__lte=last_day_of_year
    # ).aggregate(Sum('montant'))['montant__sum'] or Decimal('0')

    # total_sorties_mois = OperationSortir.objects.filter(
    #     date_de_sortie__gte=first_day_of_year,
    #     date_de_sortie__lte=last_day_of_year
    # ).aggregate(Sum('montant'))['montant__sum'] or Decimal('0')
    
    # Données des entrées par mois pour l'année sélectionnée
    entrees_par_mois = list(OperationEntrer.objects.filter(
        date_transaction__gte=first_day_of_year,
        date_transaction__lte=last_day_of_year
    ).annotate(
        mois=TruncMonth('date_transaction')
    ).values('mois').annotate(
        total=Sum('montant')
    ).order_by('mois'))

    # Données des sorties par mois pour l'année sélectionnée
    sorties_par_mois = list(OperationSortir.objects.filter(
        date_de_sortie__gte=first_day_of_year,
        date_de_sortie__lte=last_day_of_year
    ).annotate(
        mois=TruncMonth('date_de_sortie')
    ).values('mois').annotate(
        total=Sum('montant')
    ).order_by('mois'))

    # Créer un dictionnaire pour faciliter l'accès aux totaux par mois
    entrees_dict = {item['mois'].strftime('%Y-%m'): item['total'] or Decimal('0') for item in entrees_par_mois}
    sorties_dict = {item['mois'].strftime('%Y-%m'): item['total'] or Decimal('0') for item in sorties_par_mois}

    # Obtenir tous les mois uniques
    all_months = sorted(set(entrees_dict.keys()) | set(sorties_dict.keys()))

    # Calculer les soldes cumulatifs
    solde_cumule = Decimal('0')
    soldes_par_mois = []
    formatted_entrees = []
    formatted_sorties = []
    
    # Calculer le solde initial
    solde_initial = (
        OperationEntrer.objects.filter(date_transaction__lt=first_day_of_year).aggregate(Sum('montant'))['montant__sum'] or Decimal('0')
    ) - (
        OperationSortir.objects.filter(date_de_sortie__lt=first_day_of_year).aggregate(Sum('montant'))['montant__sum'] or Decimal('0')
    )
    
    solde_cumule = solde_initial

    for mois_str in all_months:
        entree_mois = entrees_dict.get(mois_str, Decimal('0'))
        sortie_mois = sorties_dict.get(mois_str, Decimal('0'))
        solde_mois = entree_mois - sortie_mois
        solde_cumule += solde_mois
        
        # Convertir la chaîne de date en objet datetime pour le formatage
        mois_date = datetime.strptime(mois_str, '%Y-%m')
        mois_format = format_date(mois_date, format='MMMM yyyy', locale='fr_FR')
        
        soldes_par_mois.append({
            'mois': mois_format,
            'solde': float(solde_cumule)  # Convertir en float pour JSON
        })
        formatted_entrees.append({
            'mois': mois_format,
            'total': float(entree_mois)  # Convertir en float pour JSON
        })
        formatted_sorties.append({
            'mois': mois_format,
            'total': float(sortie_mois)  # Convertir en float pour JSON
        })

    # Données pour le graphique des catégories de sorties
    sorties_categories = list(OperationSortir.objects.filter(
        date_de_sortie__gte=first_day_of_year,
        date_de_sortie__lte=last_day_of_year
    ).values(
        'categorie__name'
    ).annotate(
        total=Sum('montant')
    ).order_by('-total'))
    

    # Données pour le graphique des catégories d'entrées
    entrees_categories = list(OperationEntrer.objects.filter(
        date_transaction__gte=first_day_of_year,
        date_transaction__lte=last_day_of_year
    ).values(
        'categorie__name'
    ).annotate(
        total=Sum('montant')
    ).order_by('-total')[:5])

    # Formater les données des sorties par catégories
    formatted_sorties_categories = [{
        'categorie': item['categorie__name'],
        'total': float(item['total'] or Decimal('0'))
    } for item in sorties_categories]
    
    # Formater les données des entrées par catégories
    formatted_entrees_categories = [{
        'categorie': item['categorie__name'],
        'total': float(item['total'] or Decimal('0'))
    } for item in entrees_categories]

    # Calculer les totaux pour le contexte
    # total_entrees = sum(entrees_dict.values(), Decimal('0'))
    # total_sorties = sum(sorties_dict.values(), Decimal('0'))

    # Liste des années disponibles pour le formulaire de sélection
    years = range(today.year - 5, today.year + 1)  # Par exemple, les 5 dernières années

    # Formater les données pour le template
    context = {
        'solde_actuel': float(solde_cumule),
        'total_entrees': float(total_entrees_mois),  # Total des entrées de l'année sélectionnée
        'total_sorties': float(total_sorties_mois),  # Total des sorties de l'année sélectionnée
        'entrees_par_mois': json.dumps(formatted_entrees[::-1]),
        'sorties_par_mois': json.dumps(formatted_sorties[::-1]),
        'soldes_par_mois': json.dumps(soldes_par_mois),
        'sorties_categories': json.dumps(formatted_sorties_categories),
        'entrees_categories': json.dumps(formatted_entrees_categories),
        'entrees_4_mois': json.dumps(formatted_entrees[-4:][::-1]) if formatted_entrees else json.dumps([]),
        'sorties_4_mois': json.dumps(formatted_sorties[-4:][::-1]) if formatted_sorties else json.dumps([]),
        'years': years,
        'selected_year': selected_year,
        'depense_month': depense_month,
    }

    return render(request, "caisse/dashboard.html", context)



@login_required(login_url="accounts:login_user")
def operations(request):
    """
    Affiche la page des opérations.
    """
    categories = Categorie.objects.all()
    categories_entree = Categorie.objects.filter(type='entree')
    categories_sortie = Categorie.objects.filter(type='sortie')
    beneficiaires = Beneficiaire.objects.all()
    fournisseurs = Fournisseur.objects.all()

    return render(request, "caisse/operations/entre-sortie.html", {
        'categories': categories,
        'categories_entree': categories_entree,
        'categories_sortie': categories_sortie,
        'beneficiaires': beneficiaires,
        'fournisseurs': fournisseurs,
        })

@login_required(login_url="accounts:login_user")
def categories(request):
    """
    Affiche la liste des catégories.
    """
    categories = Categorie.objects.all()
    categories_json = json.loads(serialize('json', categories))
    
    context = {
        'categories': json.dumps([{**item['fields'], 'id': item['pk']} for item in categories_json]),
    }
    return render(request, "caisse/categories/categories.html", context)

@login_required(login_url="accounts:login_user")
def listes(request):
    """
    Affiche la liste des opérations avec filtrage et tri.
    """
    # Récupérer les filtres de recherche et de triage
    query = request.GET.get('q') if request.GET.get('q') != "None" else None
    categorie_name = request.GET.get('categorie') if request.GET.get('categorie') != "None" else None
    beneficiaire_id = request.GET.get('beneficiaire') if request.GET.get('beneficiaire') != "None" else None
    fournisseur_id = request.GET.get('fournisseur') if request.GET.get('fournisseur') != "None" else None
    mois = request.GET.get('mois')  # Récupérer le mois sélectionné
    sort_by = request.GET.get('sort', 'date')
    ordre = request.GET.get('order', 'desc')

    # Liste des mois pour le filtrage
    mois_liste = [
        {'value': 1, 'label': 'Janvier'},
        {'value': 2, 'label': 'Février'},
        {'value': 3, 'label': 'Mars'},
        {'value': 4, 'label': 'Avril'},
        {'value': 5, 'label': 'Mai'},
        {'value': 6, 'label': 'Juin'},
        {'value': 7, 'label': 'Juillet'},
        {'value': 8, 'label': 'Août'},
        {'value': 9, 'label': 'Septembre'},
        {'value': 10, 'label': 'Octobre'},
        {'value': 11, 'label': 'Novembre'},
        {'value': 12, 'label': 'Décembre'},
    ]

    # Initialiser les queryset avec tri par défaut
    entree = OperationEntrer.objects.all().order_by('-date_transaction')
    sortie = OperationSortir.objects.all().order_by('-date_de_sortie')
    categories = Categorie.objects.all().values('name').distinct()

    # Appliquer les filtres de recherche
    if query:
        entree = entree.filter(
            Q(description__icontains=query) |
            Q(categorie__name__icontains=query) |
            Q(montant__icontains=query)
        )
        sortie = sortie.filter(
            Q(description__icontains=query) |
            Q(categorie__name__icontains=query) |
            Q(montant__icontains=query)
        )

    # Filtre par catégorie
    if categorie_name:  # Vérifier que c'est un nombre
        if Categorie.objects.all().filter(name=categorie_name, type='entree').exists():
            entree = entree.filter(categorie_id = int( Categorie.objects.all().filter(name=categorie_name, type='entree').get().id))
        else:
            entree = []
        if Categorie.objects.all().filter(name=categorie_name, type='sortie').exists():
            sortie = sortie.filter(categorie_id = int(Categorie.objects.all().filter(name=categorie_name, type='sortie').get().id))
        else:
            sortie = []

    # Filtre par mois
    if mois and mois.isdigit():  # Vérifier que c'est un nombre
        if entree:
            entree = entree.filter(date_transaction__month=int(mois))
        if sortie:
            sortie = sortie.filter(date_de_sortie__month=int(mois))

    # Filtre par bénéficiaire
    if beneficiaire_id and beneficiaire_id.isdigit():  # Vérifier que c'est un nombre
        if sortie:
            sortie = sortie.filter(beneficiaire_id=beneficiaire_id)
        if entree:
            entree = entree.filter(beneficiaire_id=beneficiaire_id)

    # Filtre par fournisseur (uniquement pour les sorties)
    if fournisseur_id and fournisseur_id.isdigit():  # Vérifier que c'est un nombre
        if sortie:
            sortie = sortie.filter(fournisseur_id=fournisseur_id)
        entree = []

    # Pagination
    lignes_par_page = str(request.GET.get('lignes', 10)) # Valeur par défaut : 10
    
    # Triage conditionnel
    if sort_by == 'categorie':
        # Trier les opérations par catégorie
        operations = sorted(
            chain(entree, sortie),
            key=lambda x: getattr(x, 'categorie').name if x.categorie else '',
            reverse=(ordre == 'desc')
        )
    elif sort_by == 'beneficiaire':
        # Trier les opérations par bénéficiaire
        operations = sorted(
            chain(entree, sortie),
            key=lambda x: str(x.beneficiaire),
            reverse=(ordre == 'desc')
        )
    elif sort_by == 'fournisseur':
        # Trier les opérations de sortie par fournisseur
        operations = sorted(
            sortie,
            key=lambda x: getattr(x, 'fournisseur').name if x.fournisseur else '',
            reverse=(ordre == 'desc')
        )
    elif sort_by == 'description':
        # Trier les opérations par description
        operations = sorted(
            chain(entree, sortie),
            key=lambda x: getattr(x, 'description'),
            reverse=(ordre == 'desc')
        )
    elif sort_by == 'quantite':
        # Trier les opérations de sortie par quantité
        operations = sorted(
            sortie,
            key=lambda x: getattr(x, 'quantite', 0),
            reverse=(ordre == 'desc')
        )
    elif sort_by == 'montant':
        # Trier les opérations par montant
        operations = sorted(
            chain(entree, sortie),
            key=lambda x: getattr(x, 'montant'),
            reverse=(ordre == 'desc')
        )
    else:
        # Trier les opérations par date
        operations = sorted(
            chain(entree, sortie),
            key=lambda x: getattr(x, 'date_transaction', None) or getattr(x, 'date_de_sortie', None),
            reverse=(ordre == 'desc')
        )

    paginator = Paginator(operations, lignes_par_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Contexte à passer au template
    context = {
        'page_obj': page_obj,
        'categories': categories,
        'beneficiaires': Beneficiaire.objects.all(),
        'fournisseurs': Fournisseur.objects.all(),
        'lignes_par_page': lignes_par_page,
        'query': query,
        'categorie_name': categorie_name,
        'beneficiaire_id': beneficiaire_id,
        'fournisseur_id': fournisseur_id,
        'mois_liste': mois_liste,
        'mois': mois,
    }
    return render(request, 'caisse/listes/listes_operations.html', context)

@login_required(login_url="accounts:login_user")
def depenses(request: WSGIRequest):
    """
    Affiche la page des dépenses avec les opérations par employé et par catégorie.
    """
    # Récupérer le mois sélectionné
    mois_selectionne = request.GET.get('mois', timezone.now().strftime('%Y-%m'))
    beneficiaire_id = request.GET.get('employe')
    categorie_id = request.GET.get('categorie_emp')
    order_in = request.GET.get('order')
    sort_by = request.GET.get('sort')

    try:
        date_debut = timezone.datetime.strptime(f"{mois_selectionne}-01", '%Y-%m-%d')
        date_fin = (date_debut + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(days=1)
    except ValueError:
        date_debut = timezone.now().replace(day=1)
        date_fin = (date_debut + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(days=1)

    # Filtrer les opérations par mois
    operations = OperationSortir.objects.filter(
        date_de_sortie__gte=date_debut,
        date_de_sortie__lte=date_fin
    )

    # Dépenses par employé
    depenses_par_employe = operations.values(
        'beneficiaire'
    ).annotate(
        total_depenses=Sum('montant'),
        nombre_depenses=Count('id'),
        categorie_plus_depensee=models.F('categorie__name')
    ).order_by('-total_depenses')

    # Filtrer les dépenses par employé
    if beneficiaire_id and beneficiaire_id.isdigit():
        depenses_par_employe = depenses_par_employe.filter(beneficiaire_id=beneficiaire_id)
    # Filtrer les dépenses par catégorie
    if categorie_id and categorie_id.isdigit():
        depenses_par_employe = depenses_par_employe.filter(categorie_id=categorie_id)
    
    for index in range(len(depenses_par_employe)):
        depenses_par_employe[index]['beneficiaire'] = Beneficiaire.objects.get(id=int(depenses_par_employe[index]['beneficiaire']))

    # Dépenses par catégorie
    depenses_par_categorie = operations.values(
        'categorie__name'
    ).annotate(
        total_depenses=Sum('montant'),
        nombre_depenses=Count('id')
    ).order_by('-total_depenses')

    employes = sorted(Beneficiaire.objects.all().distinct(), key=lambda x: str(x))
    categories = operations.values('categorie__name', 'categorie__id').distinct().order_by('categorie__name')

    # Couleurs pour les catégories
    colors = [
        '#3B82F6', '#EF4444', '#F59E0B', '#10B981', '#6366F1',
        '#EC4899', '#8B5CF6', '#14B8A6', '#F97316', '#06B6D4'
    ]
    
    for i, depense in enumerate(depenses_par_categorie):
        depense['color'] = colors[i % len(colors)]

    # Dépenses par année
    depenses_par_annee = OperationSortir.objects.annotate(
        year=TruncYear('date_de_sortie')
    ).values('year').annotate(
        total_depenses=Sum('montant')
    ).order_by('year')

    # Dictionnaire de tri pour les dépenses par employé
    employe_sort_options = {
        'employe': lambda x: (str(x['beneficiaire']) if x['beneficiaire'] else ''),
        'total_emp': lambda x: x['total_depenses'],
        'operation_emp': lambda x: x['nombre_depenses'],
        'categorie_emp': lambda x: x['categorie_plus_depensee']
    }

    # Dictionnaire de tri pour les dépenses par catégorie
    categorie_sort_options = {
        'categorie_cat': lambda x: x['categorie__name'],
        'total_cat': lambda x: x['total_depenses'],
        'operation_cat': lambda x: x['nombre_depenses']
    }

    # Trier les dépenses par employé
    if sort_by in employe_sort_options:
        depenses_par_employe = sorted(depenses_par_employe, key=employe_sort_options[sort_by], reverse=(order_in == 'desc'))

    # Trier les dépenses par catégorie
    if sort_by in categorie_sort_options:
        depenses_par_categorie = sorted(depenses_par_categorie, key=categorie_sort_options[sort_by], reverse=(order_in == 'desc'))

    # Pagination
    lignes_par_page = request.GET.get('lignes', str(10)) # Valeur par défaut : 10

    paginator = Paginator(depenses_par_employe, lignes_par_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Générer la liste des mois
    mois_liste = []
    date_courante = timezone.now()
    for i in range(12):
        date = date_courante - timezone.timedelta(days=30*i)
        mois_liste.append({
            'value': date.strftime('%Y-%m'),
            'label': format_date(date, format='MMMM yyyy', locale='fr_FR')
        })

    context = {
        'page_obj': page_obj,
        'lignes_par_page': lignes_par_page,
        'employes': employes,
        'categories': categories,
        'depenses_par_employe': depenses_par_employe,
        'depenses_par_categorie': depenses_par_categorie,
        'depenses_par_annee': depenses_par_annee,
        'mois_selectionne': mois_selectionne,
        'mois_liste': mois_liste,
    }
    return render(request, "caisse/depenses/depense.html", context)

# Gestion des acteurs

@login_required(login_url="accounts:login_user")
def acteurs(request):
    """
    Affiche la page des acteurs (personnels, fournisseurs, catégories).
    """
    personnels = Personnel.objects.all()
    fournisseurs = Fournisseur.objects.all()
    categories = Categorie.objects.all()
    
    personnels_json = json.loads(serialize('json', personnels))
    fournisseurs_json = json.loads(serialize('json', fournisseurs))
    categories_entrees = json.loads(serialize('json', categories.filter(type='entree')))
    categories_sorties = json.loads(serialize('json', categories.filter(type='sortie')))
    
    context = {
        'personnels': json.dumps([{**item['fields'], 'id': item['pk']} for item in personnels_json]),
        'fournisseurs': json.dumps([{**item['fields'], 'id': item['pk']} for item in fournisseurs_json]),
        'categories_entrees': json.dumps([{**item['fields'], 'id': item['pk']} for item in categories_entrees]),
        'categories_sorties': json.dumps([{**item['fields'], 'id': item['pk']} for item in categories_sorties]),
        'categorie_entrees': categories.filter(type='entree'),
        'categorie_sorties': categories.filter(type='sortie')
    }
    return render(request, "caisse/acteurs/acteurs.html", context)

@login_required(login_url="accounts:login_user")
def ajouter_acteur(request: WSGIRequest):
    """
    Ajoute un nouvel acteur (fournisseur, employé ou catégorie).
    """
    data = {}
    if request.method == 'POST':
        type_acteur = request.POST.get('type_acteur')
        if type_acteur == 'fournisseurs':
            form = FournisseurForm(request.POST)
        elif type_acteur == 'employes':
            form = PersonnelForm(request.POST, request.FILES)
        elif type_acteur == 'categories':
            form = CategorieForm(request.POST)
        else:
            messages.error(request, "Type d'acteur non valide.")
            return redirect('caisse:acteurs')

        try:
            form.save()
            # Enregistrement de l'activité
            UserActivity.objects.create(user=request.user, action='Création', description=f'a ajouté un {type_acteur[:-1]}')
            messages.success(request, f"{type_acteur[:-1].capitalize()} ajouté avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'ajout de l'acteur: {str(e)}")
    return redirect('caisse:acteurs')

@login_required(login_url="accounts:login_user")
def ajouter_fournisseur(request):
    """
    Ajoute un nouveau fournisseur.
    """
    if request.method == 'POST':
        form = FournisseurForm(request.POST)
        if form.is_valid():
            form.save()
            # Enregistrement de l'activité
            UserActivity.objects.create(user=request.user, action='Création', description='a ajout un fournisseur')
            messages.success(request, "Fournisseur ajouté avec succès.")
            return redirect('caisse:acteurs')
        else:
            messages.error(request, "Erreur dans le formulaire. Veuillez vérifier les données.")
    return redirect('caisse:acteurs')

@login_required(login_url="accounts:login_user")
@require_POST
@csrf_exempt
def modifier_acteur(request, type_acteur, pk):
    if type_acteur == 'personnel':
        acteur = get_object_or_404(Personnel, pk=pk)
    elif type_acteur == 'fournisseur':
        acteur = get_object_or_404(Fournisseur, pk=pk)
    else:
        return JsonResponse({'success': False, 'error': 'Type d\'acteur invalide'}, status=400)

    data = json.loads(request.body)
    
    for key, value in data.items():
        setattr(acteur, key, value)
    
    acteur.save()
    # Enregistrement de l'activité
    UserActivity.objects.create(user=request.user, action='Modification', description='a modifier un acteur')
    return JsonResponse({
        'success': True,
        'acteur': {
            'id': acteur.id,
            'name': acteur.name if hasattr(acteur, 'name') else f"{acteur.first_name} {acteur.last_name}",
            'contact': acteur.contact if hasattr(acteur, 'contact') else acteur.email,
        }
    })

@login_required(login_url="accounts:login_user")
@require_POST
@csrf_exempt
def supprimer_acteur(request, type_acteur, pk):
    if type_acteur == 'personnel':
        acteur = get_object_or_404(Personnel, pk=pk)
    elif type_acteur == 'fournisseur':
        acteur = get_object_or_404(Fournisseur, pk=pk)
    else:
        messages.error(request, "Type d'acteur non valide.")
        return JsonResponse({'success': False, 'error': 'Type d\'acteur invalide'}, status=400)

    try:
        acteur.delete()
        UserActivity.objects.create(user=request.user, action='Suppression', description='a supprimé un acteur')
        messages.success(request, "Acteur supprimé avec succès.")
        return JsonResponse({'success': True})
    except django.db.models.deletion.ProtectedError:
        messages.error(request, "Impossible de supprimer un acteur utilisé dans des opérations.")
        return JsonResponse({'success': False, 'error': 'Impossible de supprimer un acteur utilisé dans des opérations.'}, status=400)
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression de l'acteur.{e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

# Gestion des catégories

@login_required(login_url="accounts:login_user")
def ajouter_categorie(request):
    """
    Ajoute une nouvelle catégorie.
    """
    if request.method == 'POST':
        form = CategorieForm(request.POST)
        if form.is_valid():
            form.save()
            # Enregistrement de l'activité
            UserActivity.objects.create(user=request.user, action='Création', description='a ajouté une catégorie')
            messages.success(request, "Catégorie ajoutée avec succès.")
            return redirect('caisse:acteurs')
    return redirect('caisse:acteurs')

@login_required(login_url="accounts:login_user")
@require_POST
@csrf_exempt
def modifier_categorie(request, pk):
    categorie = get_object_or_404(Categorie, pk=pk)
    data = json.loads(request.body)
    
    categorie.name = data.get('name', categorie.name)
    categorie.description = data.get('description', categorie.description)
    categorie.type = data.get('type', categorie.type)
    
    categorie.save()
    # Enregistrement de l'activité
    UserActivity.objects.create(user=request.user, action='Modification', description='a modifier une catégorie')
    return JsonResponse({
        'success': True,
        'categorie': {
            'id': categorie.id,
            'name': categorie.name,
            'description': categorie.description,
            'type': categorie.type
        }
    })

@login_required(login_url="accounts:login_user")
def supprimer_categorie(request, pk: int):
    categorie = get_object_or_404(Categorie, pk=pk)

    try:
        categorie.delete()
        UserActivity.objects.create(user=request.user, action='Suppression', description='a supprimé une catégorie')
        messages.success(request, "Catégorie supprimée avec succès.")
    except django.db.models.deletion.ProtectedError:
        messages.error(request, "Impossible de supprimer une catégorie utilisée dans des opérations.")
    except Exception:
        messages.error(request, f"Erreur lors de la suppression.")
    return redirect('caisse:acteurs')

# Gestion des opérations

@login_required(login_url="accounts:login_user")
def ajouts_entree(request):
    """
    Gère l'ajout d'opérations d'entrée.
    """
    categories_entree = Categorie.objects.filter(type='entree')
    beneficiaires = Beneficiaire.objects.all()
    
    if request.method == 'POST':
        dates = request.POST.getlist('date')
        designations = request.POST.getlist('designation')
        montants = request.POST.getlist('montant')
        categories_ids = request.POST.getlist('categorie')
        beneficiaires_ids = request.POST.get('beneficiaire-entree')
        clients = request.POST.getlist('client')

        # Vérifier la cohérence des données
        if not all([dates, designations, montants, categories_ids]):
            messages.error(request, "Veuillez remplir toutes les lignes du formulaire.")
            return render(request, 'caisse/operations/entre-sortie.html', {
                'categories_entree': categories_entree,
                'beneficiaires': beneficiaires,
                'operation': 'entree',
            })

        # try:
        for i in range(len(dates)):
            # Validation des données individuelles
            if not dates[i] or not designations[i] or not montants[i] or not categories_ids[i]:
                messages.error(request, f"Ligne {i+1} : Des champs obligatoires sont manquants.")
            
            montant = float(montants[i])
            if montant < 0:
                messages.error(request, f"Ligne {i+1} : Le montant doit être positif.")

            try:
                OperationEntrer.objects.create(
                    date_transaction=parse_date(dates[i]),
                    description=designations[i],
                    montant=montant,
                    categorie_id=int(categories_ids[i]),
                    beneficiaire_id=int(beneficiaires_ids) if beneficiaires_ids else int(Beneficiaire.objects.get(name="Général").id),
                    client=clients[i] if clients else ""
                )
                messages.success(request, f"Ligne {i+1} : Les opérations d'entrée ont été ajoutées avec succès.")
            except:
                messages.error(request, f"Ligne {i+1} : Erreur lors de l'ajout de l'opération.")
                continue

    return render(request, 'caisse/operations/entre-sortie.html', {
        'categories_entree': categories_entree,
        'beneficiaires': beneficiaires,
        'operation': 'entree',
    })

@login_required(login_url="accounts:login_user")
def ajouts_sortie(request):
    """
    Gère l'ajout d'opérations de sortie sans vérification.
    """
    categories_sortie = Categorie.objects.filter(type='sortie') 
    beneficiaires = Beneficiaire.objects.all()
    fournisseurs = Fournisseur.objects.all()

    if request.method == 'POST':
        dates = request.POST.getlist('date')
        designations = request.POST.getlist('designation')
        beneficiaires_ids = request.POST.getlist('beneficiaire-sortie')
        fournisseurs_ids = request.POST.getlist('fournisseur')
        quantites = request.POST.getlist('quantite')
        prix_unitaires = request.POST.getlist('prixUnitaire')
        categories_ids = request.POST.getlist('categorie')

        for i in range(len(dates)):
            try:
                # Validation de base des données
                if not dates[i] or not designations[i] :
                    raise ValueError("Tous les champs doivent être remplis.")

                date_operation = datetime.strptime(dates[i], '%Y-%m-%d').date()
                quantite = int(quantites[i])
                prix_unitaire = float(prix_unitaires[i])

                if quantite <= 0 or prix_unitaire < 0:
                    raise ValueError("Quantité et prix unitaire doivent être positifs.")

                # Création de l'opération
                OperationSortir.objects.create(
                    date_de_sortie=date_operation,
                    description=designations[i],
                    quantite=quantite,
                    montant=quantite * prix_unitaire,
                    categorie_id=int(categories_ids[i]),
                    beneficiaire_id=int(beneficiaires_ids[i]) if beneficiaires_ids else int(Beneficiaire.objects.get(name="Général").id),
                    fournisseur_id=int(fournisseurs_ids[i])
                )

            except Exception as e:
                messages.error(request, f"Erreur à la ligne {i + 1} : {e}")
                return render(request, 'caisse/operations/entre-sortie.html', {
                    'categories_sortie': categories_sortie,
                    'beneficiaires': beneficiaires,
                    'fournisseurs': fournisseurs,
                    'operation': 'sortie',
                })

        # Ajout des opérations réussi
        messages.success(request, "Les opérations de sortie ont été ajoutées avec succès.")
        return redirect('caisse:liste_sorties')

    return render(request, 'caisse/operations/entre-sortie.html', {
        'categories_sortie': categories_sortie,
        'beneficiaires': beneficiaires,
        'fournisseurs': fournisseurs,
        'operation': 'sortie',
    })

@login_required(login_url="accounts:login_user")
def modifier_entree(request, pk):
    # Récupérer l'entrée existante
    entree = get_object_or_404(OperationEntrer, id=pk)

    if request.method == 'POST':
        # Récupérer les données du formulaire
        date_transaction = request.POST.get('date')
        description = request.POST.get('description')
        montant = request.POST.get('montant')
        beneficiaire_id = request.POST.get('beneficiaire')
        categorie_id = request.POST.get('categorie')
        client = request.POST.get('client')

        # Mettre à jour l'objet entrée
        entree.date_transaction = date_transaction
        entree.description = description
        entree.montant = montant
        entree.beneficiaire_id = beneficiaire_id
        entree.categorie_id = categorie_id
        entree.client = client

        # Sauvegarder les modifications
        entree.save()
        
        # Enregistrement de l'activité
        UserActivity.objects.create(user=request.user, action='Modification', description='a modifier une opération entrée')
        # Ajouter un message de succès
        messages.success(request, "L'opération d'entrée a été modifiée avec succès.")

        # Rediriger vers la liste des entrées
        return redirect(reverse('caisse:liste_entrees'))

    # Préparer le contexte pour le template
    context = {
        'entree': entree,
        'categories_entree': Categorie.objects.filter(type='entree'),
        'beneficiaires': Beneficiaire.objects.all(),
    }

    # Rendre le template avec le contexte
    return render(request, 'caisse/listes/modifier_entree.html', context)


@login_required(login_url="accounts:login_user")
def modifier_sortie(request, pk):
    # Récupérer l'opération de sortie spécifique
    operation = get_object_or_404(OperationSortir, id=pk)

    if request.method == 'POST':
        # Récupérer les données du formulaire
        date_de_sortie = request.POST.get('date')
        designation = request.POST.get('designation')
        quantite = request.POST.get('quantite')
        prix_unitaire = request.POST.get('prixUnitaire')
        beneficiaire_id = request.POST.get('beneficiaire')
        fournisseur_id = request.POST.get('fournisseur')
        categorie_id = request.POST.get('categorie')

        # Mettre à jour les champs de l'opération
        operation.date_de_sortie = date_de_sortie
        operation.description = designation
        operation.quantite = quantite
        operation.montant = prix_unitaire  # En supposant que 'montant' correspond au prix unitaire
        operation.beneficiaire_id = beneficiaire_id
        operation.fournisseur_id = fournisseur_id
        operation.categorie_id = categorie_id

        # Sauvegarder les modifications
        operation.save()
        # Enregistrement de l'activité
        UserActivity.objects.create(user=request.user, action='Modification', description='a modifier une opération sortie')
        # Ajouter un message de succès
        messages.success(request, "L'opération a été modifiée avec succès.")
        # Rediriger vers la liste des sorties
        return redirect(reverse('caisse:liste_sorties'))

    # Préparer le contexte pour le template avec les options de sélection
    context = {
        'operation': operation,
        'beneficiaires': Beneficiaire.objects.all(),
        'fournisseurs': Fournisseur.objects.all(),
        'categories_sortie': Categorie.objects.filter(type='sortie'),
    }

    # Rendre le template avec le contexte
    return render(request, 'caisse/listes/modifier_sortie.html', context)

#Suppression des entrées
@login_required(login_url="accounts:login_user")
def supprimer_entree(request, pk):
    """
    Supprime une opération d'entrée ou de sortie en fonction de son ID.
    """

    operation = OperationEntrer.objects.get(pk=pk)
    
    operation.delete()
    UserActivity.objects.create(user=request.user, action='Suppression', description='a supprimé une opération entrée')
    messages.success(request, "L'opération a été supprimée avec succès.")

    ligne = request.GET.get('lignes')
    
    return redirect(request.META.get('HTTP_REFERER'))

#Suppression des sorties
@login_required(login_url="accounts:login_user")
def supprimer_sortie(request, pk):
    """
    Supprime une opération d'entrée ou de sortie en fonction de son ID.
    """

    operation = OperationSortir.objects.get(pk=pk)
    
    operation.delete()
    UserActivity.objects.create(user=request.user, action='Suprression', description='a supprimé une opération sortie')
    messages.success(request, "L'opération a été supprimée avec succès.")
    return redirect(request.META.get('HTTP_REFERER'))

# Add this new view
@login_required(login_url="accounts:login_user")
def parametres(request):
    """
    Affiche la page des paramètres.
    """
    context = {
        # You can add any necessary context data here
    }
    return render(request, "caisse/parametres/parametres.html", context)

@login_required(login_url="accounts:login_user")
@require_POST
@csrf_exempt
def creer_categorie(request: WSGIRequest):
    query_string = request.body

    # Décoder la chaîne de requête
    decoded_string = query_string.decode('utf-8')
    data = urllib.parse.parse_qs(decoded_string)

    name = data.get('categorie_name')[0]
    description = data.get('categorie_description', [None])[0]
    type = data.get('categorie_type')[0]
    
    # Vérifie si le nom et le type sont renseignés
    if not name or not type:
        messages.error(request, f"Erreur : Le nom et le type sont obligatoires.")
        return redirect('caisse:acteurs')

    # Vérifier si la catégorie existe déjà par nom et type
    if Categorie.objects.filter(name=name, type=type).exists():
        messages.error(request, "Une catégorie avec ce nom et ce type existe déjà.")
        return redirect('caisse:acteurs')
    
    try:
        Categorie.objects.create(name=name, description=description, type=type)    
        messages.success(request, "Catégorie créée avec succès.")

        UserActivity.objects.create(user=request.user, action='Création', description='a créé une nouvelle catégorie')
    except Exception as e:
        messages.error(request, f"Erreur lors de la création de la catégorie.")
    
    return redirect('caisse:acteurs')

@login_required(login_url="accounts:login_user")
def editer_acteur(request, type_acteur, pk):
    if type_acteur == 'personnel':
        acteur = get_object_or_404(Personnel, pk=pk)
        form_class = PersonnelForm
    elif type_acteur == 'fournisseur':
        acteur = get_object_or_404(Fournisseur, pk=pk)
        form_class = FournisseurForm
    else:
        messages.error(request, "Type d'acteur non valide.")
        return redirect('caisse:acteurs')

    if request.method == 'POST':
        form = form_class(request.POST, request.FILES, instance=acteur)
        if form.is_valid():
            form.save()
            UserActivity.objects.create(user=request.user, action='Modification', description=f'a modifié un {type_acteur}')
            messages.success(request, f"{type_acteur.capitalize()} modifié avec succès.")
            return redirect('caisse:acteurs')
    else:
        form = form_class(instance=acteur)

    context = {
        'form': form,
        'type_acteur': type_acteur,
        'acteur': acteur,
    }
    return render(request, 'caisse/acteurs/editer_acteur.html', context)

@login_required(login_url="accounts:login_user")
def editer_categorie(request, pk):
    categorie = get_object_or_404(Categorie, pk=pk)
    if request.method == 'POST':
        form = CategorieForm(request.POST, instance=categorie)
        if form.is_valid():
            form.save()
            messages.success(request, "Catégorie modifiée avec succès.")
            return redirect('caisse:acteurs')
    else:
        form = CategorieForm(instance=categorie)
    
    context = {
        'form': form,
        'categorie': categorie,
    }
    return render(request, 'caisse/acteurs/editer_categorie.html', context)

@login_required(login_url="accounts:login_user")
@user_passes_test(is_admin)
def utilisateurs(request):
    """
    Affiche la liste des utilisateurs.
    """
    users = User.objects.all()
    return render(request, 'caisse/utilisateurs/utilisateurs.html', {'users': users})

@login_required(login_url="accounts:login_user")
@user_passes_test(is_admin)
@csrf_exempt
def creer_utilisateur(request):
    """
    Crée un nouvel utilisateur.
    """
    if request.method != 'POST':
        return redirect('caisse:utilisateurs')
    
    try:
        # Récupérer les données du formulaire
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        is_staff = request.POST.get('is_staff') == 'true'
        is_active = request.POST.get('is_active') == 'true'
        
        # Vérifier si l'utilisateur existe déjà
        if User.objects.filter(username=username).exists():
            messages.error(request, "Un utilisateur avec ce nom d'utilisateur existe déjà")
            return redirect('caisse:utilisateurs')
            
        if User.objects.filter(email=email).exists():
            messages.error(request, "Un utilisateur avec cet email existe déjà")
            return redirect('caisse:utilisateurs')
        
        # Créer l'utilisateur
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            is_staff=is_staff,
            is_active=is_active
        )

        # Gérer l'upload de la photo
        if 'photo' in request.FILES:
            user.photo = request.FILES['photo']
            user.save()
        
        messages.success(request, "Utilisateur créé avec succès")
        return redirect('caisse:utilisateurs')
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la création de l'utilisateur: {str(e)}")
        return redirect('caisse:utilisateurs')

@login_required(login_url="accounts:login_user")
@user_passes_test(is_admin)
@require_POST
@csrf_exempt
def modifier_utilisateur(request, pk):
    """
    Modifie un utilisateur existant.
    """
    user = get_object_or_404(User, pk=pk)
    
    try:
        # Si les données sont envoyées en multipart/form-data
        if request.content_type and 'multipart/form-data' in request.content_type:
            data = request.POST.dict()
            # Convertir les chaînes 'true'/'false' en booléens
            data['is_staff'] = data.get('is_staff', 'false').lower() == 'true'
            data['is_active'] = data.get('is_active', 'true').lower() == 'true'
        else:
            # Si les données sont envoyées en JSON
            data = json.loads(request.body)
        
        # Vérifier si le nom d'utilisateur existe déjà pour un autre utilisateur
        if User.objects.exclude(pk=pk).filter(username=data['username']).exists():
            return JsonResponse({
                'success': False, 
                'error': "Un utilisateur avec ce nom d'utilisateur existe déjà"
            }, status=400)
            
        if User.objects.exclude(pk=pk).filter(email=data['email']).exists():
            return JsonResponse({
                'success': False, 
                'error': "Un utilisateur avec cet email existe déjà"
            }, status=400)
        
        user.username = data['username']
        user.email = data['email']
        user.first_name = data.get('first_name', '')
        user.last_name = data.get('last_name', '')
        user.is_staff = data.get('is_staff', False)
        user.is_active = data.get('is_active', True)
        
        if data.get('password'):
            user.set_password(data['password'])
            
        # Gérer l'upload de la photo
        if 'photo' in request.FILES:
            user.photo = request.FILES['photo']
            
        user.save()
        # Enregistrement de l'activité
        UserActivity.objects.create(user=request.user, action='Modification', description='a modifier un utilisateur')
        return JsonResponse({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'is_staff': user.is_staff,
                'is_active': user.is_active,
                'photo_url': user.photo.url if user.photo else None
            }
        })
    except KeyError as e:
        return JsonResponse({
            'success': False, 
            'error': f"Champ requis manquant: {str(e)}"
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=400)

@login_required(login_url="accounts:login_user")
@user_passes_test(is_admin)
@require_POST
@csrf_exempt
def supprimer_utilisateur(request, pk):
    """
    Supprime un utilisateur.
    """
    user = get_object_or_404(User, pk=pk)
    
    try:
        user.delete()
        UserActivity.objects.create(user=request.user, action='Suppression', description='a supprimé un utilisateur')
        messages.success(request, f"Utilisateur {user.get_full_name()} supprimé avec succès.")
        return JsonResponse({'success': True})
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression de l'utilisateur {user.get_full_name()}.")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required(login_url="accounts:login_user")
@user_passes_test(is_admin)
def editer_utilisateur(request, pk):
    """
    Affiche le formulaire de modification d'un utilisateur.
    """
    user = get_object_or_404(User, pk=pk)
    if request.method == 'GET':
        return render(request, 'caisse/utilisateurs/modifier_utilisateur.html', {'user': user})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)

@login_required(login_url="accounts:login_user")
def update_profile(request):
    if request.method == 'POST':
        user = request.user
        
        # Mise à jour des informations de base
        user.first_name = request.POST.get('first_name', '')
        user.last_name = request.POST.get('last_name', '')
        user.email = request.POST.get('email', '')
        user.phone = request.POST.get('phone', '')

        # Gestion de la photo de profil
        if 'photo' in request.FILES:
            user.photo = request.FILES['photo']
        
        try:
            user.save()
            UserActivity.objects.create(user=request.user, action='Modification', description='a modifié son profil')
            messages.success(request, 'Votre profil a été mis à jour avec succès.')
        except Exception as e:
            messages.error(request, f'Erreur lors de la mise à jour du profil: {str(e)}')
        
        return redirect('caisse:parametres')
    
    return redirect('caisse:parametres')

@login_required(login_url="accounts:login_user")
def change_password(request):
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        user = request.user
        
        # Vérification du mot de passe actuel
        if not user.check_password(current_password):
            messages.error(request, 'Le mot de passe actuel est incorrect.')
            return redirect('caisse:parametres')
        
        # Vérification de la correspondance des nouveaux mots de passe
        if new_password != confirm_password:
            messages.error(request, 'Les nouveaux mots de passe ne correspondent pas.')
            return redirect('caisse:parametres')
        
        # Mise à jour du mot de passe
        try:
            user.set_password(new_password)
            user.save()
            UserActivity.objects.create(user=request.user, action='Modification', description='a modifié son mot de passe')
            update_session_auth_hash(request, user)  # Garde l'utilisateur connecté
            messages.success(request, 'Votre mot de passe a été modifié avec succès.')
        except Exception as e:
            messages.error(request, f'Erreur lors du changement de mot de passe: {str(e)}')
        
        return redirect('caisse:parametres')
    
    return redirect('caisse:parametres')

@login_required(login_url="accounts:login_user")
def liste_entrees(request):
    """
    Affiche la liste des opérations d'entrée.
    """
    # Récupérer les filtres de recherche et de triage
    query = request.GET.get('q')
    categorie_id = request.GET.get('categorie')
    beneficiaire_id = request.GET.get('beneficiaire')
    mois = request.GET.get('mois')
    sort_by = request.GET.get('sort', 'date')  # Trier par date par défaut
    ordre = request.GET.get('order', 'desc')  # Ordre décroissant par défaut
    
    # Liste des mois pour le filtrage
    mois_liste = [
        {'value': 1, 'label': 'Janvier'},
        {'value': 2, 'label': 'Février'},
        {'value': 3, 'label': 'Mars'},
        {'value': 4, 'label': 'Avril'},
        {'value': 5, 'label': 'Mai'},
        {'value': 6, 'label': 'Juin'},
        {'value': 7, 'label': 'Juillet'},
        {'value': 8, 'label': 'Août'},
        {'value': 9, 'label': 'Septembre'},
        {'value': 10, 'label': 'Octobre'},
        {'value': 11, 'label': 'Novembre'},
        {'value': 12, 'label': 'Décembre'},
    ]

    # Filtrer les opérations d'entrée
    entrees = OperationEntrer.objects.all()

    # Filtre par requête de recherche
    if query:
        entrees = entrees.filter(
            Q(description__icontains=query) | 
            Q(categorie__name__icontains=query) |
            Q(montant__icontains=query) | 
            Q(date_transaction__icontains=query)
        )
    
    # Filtre par catégorie
    if categorie_id and categorie_id.isdigit():
        entrees = entrees.filter(categorie_id=categorie_id)    
    # Filtre par bénéficiaire
    if beneficiaire_id and beneficiaire_id.isdigit():
        entrees = entrees.filter(beneficiaire_id=beneficiaire_id)
    # Filtre par client
    client = request.GET.get('client')
    if client:
        if client != 'Pas de client':
            entrees = entrees.filter(client=client)
        else:
            entrees = entrees.filter(client="")
    # Filtre par mois
    if mois and mois.isdigit():
        entrees = entrees.filter(date_transaction__month=int(mois))

    # Définir les champs de tri valides
    valid_sort_fields = {
        'description': 'description',
        'categorie': 'categorie__name',
        'client': 'client',
        'beneficiaire': 'beneficiaire__name',
        'date': 'date_transaction',
        'montant': 'montant'
    }

    # Appliquer le tri
    if sort_by in valid_sort_fields:
        sort_field = valid_sort_fields[sort_by]
        if ordre == 'desc':
            sort_field = f'-{sort_field}'  # Tri décroissant
        entrees = entrees.order_by(sort_field)
        
    # Récupérer uniquement les catégories de type "entrée" pour les options de filtrage
    categories = Categorie.objects.filter(type="entree")
    beneficiaires = Beneficiaire.objects.all()
    clients = OperationEntrer.objects.values_list('client', flat=True).distinct()
    clients = [client if client != '' else 'Pas de client' for client in clients]
    
    # Charger le template
    template = loader.get_template('caisse/listes/entrees.html')

    # Pagination
    lignes_par_page = str(request.GET.get('lignes', 10))  # Valeur par défaut : 10
    paginator = Paginator(entrees, lignes_par_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Contexte à passer au template
    context = {
        'page_obj': page_obj,
        'categories': categories,
        'beneficiaires': beneficiaires,
        'clients': clients,
        'prix': "Ar",
        'sort_by': sort_by,
        'ordre': ordre,
        'lignes_par_page': lignes_par_page,
        'query': query,
        'categorie_id': categorie_id,
        'beneficiaire_id': beneficiaire_id,
        'mois_liste': mois_liste, 
        'mois': mois, 
    }
    return HttpResponse(template.render(context, request))

@login_required(login_url="accounts:login_user")
def liste_sorties(request):
    """
    Affiche la liste des opérations de sortie.
    """
    # Récupérer les filtres de recherche et de triage
    query = request.GET.get('q')
    categorie_id = request.GET.get('categorie')
    beneficiaire_id = request.GET.get('beneficiaire')
    fournisseur_id = request.GET.get('fournisseur')
    mois = request.GET.get('mois')
    sort_by = request.GET.get('sort', 'date')  # Trier par date par défaut
    ordre = request.GET.get('order', 'desc')  # Ordre décroissant par défaut
    
    # Liste des mois pour le filtrage
    mois_liste = [
        {'value': 1, 'label': 'Janvier'},
        {'value': 2, 'label': 'Février'},
        {'value': 3, 'label': 'Mars'},
        {'value': 4, 'label': 'Avril'},
        {'value': 5, 'label': 'Mai'},
        {'value': 6, 'label': 'Juin'},
        {'value': 7, 'label': 'Juillet'},
        {'value': 8, 'label': 'Août'},
        {'value': 9, 'label': 'Septembre'},
        {'value': 10, 'label': 'Octobre'},
        {'value': 11, 'label': 'Novembre'},
        {'value': 12, 'label': 'Décembre'},
    ]

    # Filtrer les opérations de sortie
    sorties = OperationSortir.objects.all()

    if query:
        sorties = sorties.filter(
            Q(description__icontains=query) | 
            Q(categorie__name__icontains=query) |
            Q(montant__icontains=query) | 
            Q(date_de_sortie__icontains=query)
        )
    if categorie_id and categorie_id.isdigit():
        sorties = sorties.filter(categorie_id=categorie_id)
    if beneficiaire_id and beneficiaire_id.isdigit():
        sorties = sorties.filter(beneficiaire_id=beneficiaire_id)
    if fournisseur_id and fournisseur_id.isdigit():
        sorties = sorties.filter(fournisseur_id=fournisseur_id)
    # Filtre par mois
    if mois and mois.isdigit():  # Vérifiez que mois est un nombre
        sorties = sorties.filter(date_de_sortie__month=int(mois))

    # Définir les champs de tri valides
    valid_sort_fields = {
        'description': 'description',
        'categorie': 'categorie__name',
        'date': 'date_de_sortie',
        'beneficiaire': 'beneficiaire__name',
        'fournisseur': 'fournisseur__name',
        'montant': 'montant',
        'quantite': 'quantite'
    }

    # Appliquer le tri
    if sort_by in valid_sort_fields:
        sort_field = valid_sort_fields[sort_by]
        if ordre == 'desc':
            sort_field = f'-{sort_field}'
        sorties = sorties.order_by(sort_field)

    # Récupérer uniquement les catégories de type "sortie" pour les options de filtrage
    categories = Categorie.objects.filter(type="sortie")
    beneficiaires = Beneficiaire.objects.all()
    fournisseurs = Fournisseur.objects.all()
    
    # Charger le template
    template = loader.get_template('caisse/listes/sorties.html')

    # Pagination
    lignes_par_page = str(request.GET.get('lignes', 10))  # Valeur par défaut : 10
    paginator = Paginator(sorties, lignes_par_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Contexte à passer au template
    context = {
        'page_obj': page_obj,
        'categories': categories,
        'beneficiaires': beneficiaires,
        'fournisseurs': fournisseurs,
        'prix': "Ar",
        'sort_by': sort_by,
        'ordre': ordre,
        'lignes_par_page': lignes_par_page,
        'query': query,
        'categorie_id': categorie_id,
        'beneficiaire_id': beneficiaire_id,
        'fournisseur_id': fournisseur_id,
        'mois_liste': mois_liste,
        'mois': mois,
    }
    return HttpResponse(template.render(context, request))


#Pour générer un rapport en EXCEL (.xlsx)
def generer_excel_operations(request):
    # Vérifie si l'utilisateur souhaite exporter toutes les opérations ou seulement celles sélectionnées
    if request.POST.get("export_all"):
        operations_entrer = OperationEntrer.objects.all()
        operations_sortir = OperationSortir.objects.all()
    else:
        selected_ids = request.POST.getlist("selected_operations")
        operations_entrer = OperationEntrer.objects.filter(id__in=selected_ids)
        operations_sortir = OperationSortir.objects.filter(id__in=selected_ids)

    # Création d'un nouveau classeur Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Liste des Opérations"

    # En-têtes
    headers = ["Type", "Description", "Catégorie", "Bénéficiaire", "Fournisseur", "Date", "Quantité", "Montant"]
    sheet.append(headers)

    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Définir la largeur des colonnes
    column_widths = [10, 30, 15, 25, 20, 20, 15, 15]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    # Fonction pour ajouter des opérations au fichier Excel
    def ajouter_operations(operations, type_operation, avec_beneficiaire=False):
        for operation in operations:
            beneficiaire = operation.beneficiaire.name if avec_beneficiaire else "N/A"
            fournisseur = operation.fournisseur.name if avec_beneficiaire else "N/A"
            quantite = operation.quantite if avec_beneficiaire else "N/A"
            date_str = operation.date.strftime('%d-%m-%Y')
            row = [type_operation, operation.description, operation.categorie.name, beneficiaire, fournisseur, date_str, quantite, operation.montant]
            sheet.append(row)

    # Ajouter les opérations d'entrée et de sortie
    ajouter_operations(operations_entrer, "Entrée")
    ajouter_operations(operations_sortir, "Sortie", avec_beneficiaire=True)

    # Nom du fichier avec la date et l'heure actuelles
    now = datetime.now().strftime('%d-%m-%Y_%H-%M')
    filename = f"rapport_operations_{now}.xlsx"

    # Préparer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    # Enregistrer l'activité de l'utilisateur
    UserActivity.objects.create(
    user=request.user,
    action='EXPORTATION',  
    description="a exporté en Excel la liste des opérations d'entrée et de sortie"
)
    return response

def generer_excel_operations_entrees(request):
    if request.POST.get("export_all"):
        operations_entrer = OperationEntrer.objects.all()
    else:
        selected_ids = request.POST.getlist("selected_operations")
        operations_entrer = OperationEntrer.objects.filter(id__in=selected_ids)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Entrées"

    headers = ["Description", "Catégorie", "Date", "Montant"]
    sheet.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Définir la largeur des colonnes
    column_widths = [10, 30, 15, 25, 20, 20, 15, 15]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = width


    for operation in operations_entrer:
        row = [operation.description, operation.categorie.name, operation.date_transaction, operation.montant]
        sheet.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"entrees_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    # Enregistrer l'activité de l'utilisateur
    UserActivity.objects.create(
    user=request.user,
    action='EXPORTATION',  
    description="a exporté en Excel la liste des opérations d'entrée"
)

    return response

def generer_excel_operations_sorties(request):
    # Vérifier si l'utilisateur souhaite exporter toutes les opérations ou seulement celles sélectionnées
    if request.POST.get("export_all"):
        operations_sortie = OperationSortir.objects.all()
    else:
        selected_ids = request.POST.getlist("selected_operations")
        operations_sortie = OperationSortir.objects.filter(id__in=selected_ids)

    # Créer un nouveau classeur Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sorties"

    # Définir les en-têtes
    headers = ["Description", "Catégorie", "Bénéficiaire", "Fournisseur", "Date", "Quantité", "Montant"]
    sheet.append(headers)

    # Style des en-têtes
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Remplir les données des opérations de sortie
    for operation in operations_sortie:
        row = [
            operation.description,
            operation.categorie.name,
            f"{operation.beneficiaire.personnel} {operation.beneficiaire.name}",
            operation.fournisseur.name,
            operation.date_de_sortie.strftime('%d-%m-%Y'),
            operation.quantite,
            operation.montant,
        ]
        sheet.append(row)

    # Ajuster la largeur des colonnes
    column_widths = [30, 20, 25, 25, 20, 10, 15]  # Largeurs ajustées
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    # Créer la réponse HTTP pour le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"sorties_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    # Enregistrer l'activité de l'utilisateur
    UserActivity.objects.create(
        user=request.user,
        action="EXPORTATION",
        description="a exporté en Excel la liste des opérations de sortie"
    )

    return response

@login_required(login_url="accounts:login_user")
def historique(request):
    # Si l'utilisateur est un administrateur, afficher toutes les activités
    if request.user.is_staff:
        historique = UserActivity.objects.all().order_by('-timestamp')
    else:
        # Sinon, afficher uniquement les activités de l'utilisateur connecté
        historique = UserActivity.objects.filter(user=request.user).order_by('-timestamp')
    
    return render(request, 'caisse/historique/historique.html', {'historique': historique})

@login_required(login_url="accounts:login_user")
def beneficiaires(request: WSGIRequest):
    """
    Affiche la liste des bénéficiaires.
    """
    beneficiaires = Beneficiaire.objects.all()
    personnels = Personnel.objects.all()

    beneficiaire_id = request.GET.get('beneficiaire')
    type_name = request.GET.get('type')
    sort_by = request.GET.get('sort')
    order = request.GET.get('order')
    
    valeur = [(str(beneficiaire), beneficiaire.id) for beneficiaire in beneficiaires]

    # Filtrer les bénéficiaires par ID
    if beneficiaire_id and beneficiaire_id.isdigit():
        beneficiaires = beneficiaires.filter(id=beneficiaire_id)
    elif type_name:
        # Filtrer les bénéficiaires par type
        if type_name == 'Personnel interne':
            beneficiaires = beneficiaires.filter(personnel__isnull=False)
        elif type_name == 'Bénéficiaire externe':
            beneficiaires = beneficiaires.filter(personnel__isnull=True)
        else:
            messages.error(request, "Type de bénéficiaire non valide.")
    
    # Classification des opérations de chaque bénéficiaire
    operations_par_beneficiaire = defaultdict(lambda: {'entrees': 0, 'sorties': 0})
    
    for entree in OperationEntrer.objects.all():
        if entree.beneficiaire and Beneficiaire.objects.filter(id=entree.beneficiaire.id).exists():
            operations_par_beneficiaire[entree.beneficiaire.id]['entrees'] += 1
    
    for sortie in OperationSortir.objects.all():
        if sortie.beneficiaire:
            operations_par_beneficiaire[sortie.beneficiaire.id]['sorties'] += 1
    
    for beneficiaire in beneficiaires:
        beneficiaire.operations = operations_par_beneficiaire[beneficiaire.id]

    # Trier les bénéficiaires par ordre alphabétique
    if sort_by and order:
        if sort_by == "name":
            if order == 'asc':
                valeur.sort()
            elif order == 'desc':
                valeur.sort(reverse=True)
            else:
                messages.error(request, "Ordre de tri non valide.")
                return redirect('caisse:beneficiaires')
        elif sort_by == "type":
            valeur.clear()
            if order == 'desc':
                for beneficiaire in beneficiaires:
                    if beneficiaire.personnel:
                        valeur += [(str(beneficiaire), beneficiaire.id)]
                for beneficiaire in beneficiaires:
                    if not beneficiaire.personnel:
                        valeur += [(str(beneficiaire), beneficiaire.id)]
            elif order == 'asc':
                for beneficiaire in beneficiaires:
                    if beneficiaire.personnel:
                        valeur += [(str(beneficiaire), beneficiaire.id)]
                for beneficiaire in beneficiaires:
                    if not beneficiaire.personnel:
                        valeur += [(str(beneficiaire), beneficiaire.id)]
                valeur.reverse()
        elif sort_by == "operation":
            valeur.clear()
            for beneficiaire in beneficiaires:
                operations = operations_par_beneficiaire[beneficiaire.id]
                valeur += [(str(beneficiaire), beneficiaire.id, operations['entrees'], operations['sorties'], operations['entrees'] + operations['sorties'])]
            
            valeur.sort(key=lambda x: x[-1], reverse=order == 'desc')
    
    beneficiaires_sorted = sorted(beneficiaires, key=lambda b: next((i for i, v in enumerate(valeur) if v[1] == b.id), len(valeur)))

    lignes_par_page = request.GET.get('lignes', 10)

    paginator = Paginator(beneficiaires_sorted, lignes_par_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    types = ['Personnel interne', 'Bénéficiaire externe']
    
    context = {
        'beneficiaires_filtre': valeur,
        'beneficiaires': beneficiaires_sorted,
        'personnels': personnels,
        'types': types,
        'page_obj': page_obj,
        'lignes_par_page': str(lignes_par_page)
    }

    return render(request, "caisse/acteurs/beneficiaires.html", context)

@login_required(login_url="accounts:login_user")
@require_POST
@csrf_exempt
def creer_beneficiaire(request: WSGIRequest):
    query_string = request.body

    # Décoder la chaîne de requête
    decoded_string = query_string.decode('utf-8')
    data = urllib.parse.parse_qs(decoded_string)
    personnel_id = data.get('personnel_id', [None])[0]
    name = data.get('name')[0]
    
    try:
        personnel = None
        if personnel_id:
            personnel = Personnel.objects.get(pk=personnel_id)
        
        Beneficiaire.objects.create(
            personnel=personnel,
            name = name if not personnel else None
        )
        
        UserActivity.objects.create(
            user=request.user, 
            action='Création', 
            description='a créé un nouveau bénéficiaire'
        )

        messages.success(request, "Bénéficiaire créé avec succès.")
        
        return redirect('caisse:beneficiaires')
    except Exception as e:
        messages.error(request, f"Erreur lors de la création du bénéficiaire.")
        return redirect('caisse:beneficiaires') 

@login_required(login_url="accounts:login_user")
@require_POST
@csrf_exempt
def modifier_beneficiaire(request: WSGIRequest, pk):
    beneficiaire = get_object_or_404(Beneficiaire, pk=pk)

    # La chaîne de requête encodée en URL
    query_string = request.body

    # Décoder la chaîne de requête
    decoded_string = query_string.decode('utf-8')
    data = urllib.parse.parse_qs(decoded_string)
    
    try:
        personnel_id = int(data.get('personnel_id')[0])
        beneficiaire.personnel = Personnel.objects.get(id=personnel_id)
        beneficiaire.name = None

    except TypeError:
        beneficiaire.personnel = None
        beneficiaire.name = data.get('name')[0]

    except Exception as e:
        messages.error(request, f"Erreur lors de la modification du bénéficiaire.")
        return redirect(request.META.get('HTTP_REFERER'))
        
    beneficiaire.save()
    
    UserActivity.objects.create(
        user=request.user, 
        action='Modification', 
        description='a modifié un bénéficiaire'
    )

    messages.success(request, "Bénéficiaire modifié avec succès.")
    
    return redirect('caisse:beneficiaires')

@login_required(login_url="accounts:login_user")
@csrf_exempt
def supprimer_beneficiaire(request, pk):
    beneficiaire = get_object_or_404(Beneficiaire, pk=pk)
    
    try:
        beneficiaire.delete()
        UserActivity.objects.create(
            user=request.user, 
            action='Suppression', 
            description='a supprimé un bénéficiaire'
        )
        messages.success(request, "Bénéficiaire supprimé avec succès.")
    except django.db.models.deletion.ProtectedError:
        messages.error(request, "Impossible de supprimer ce bénéficiaire car il est lié à des opérations.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression du bénéficiaire.")
    
    return redirect('caisse:beneficiaires')

@login_required(login_url="accounts:login_user")
def editer_beneficiaire(request, pk):
    beneficiaire = get_object_or_404(Beneficiaire, pk=pk)
    personnels = Personnel.objects.all()
    
    context = {
        'id': int(pk),
        'beneficiaire': beneficiaire,
        'personnels': personnels,
    }
    return render(request, 'caisse/acteurs/editer_beneficiaire.html', context)

def get_available_years():
    """Récupère toutes les années disponibles dans les opérations"""
    current_year = timezone.now().year
    all_years = set()
    
    # Années des entrées
    entree_years = OperationEntrer.objects.dates('date_transaction', 'year')
    all_years.update(d.year for d in entree_years)
    
    # Années des sorties
    sortie_years = OperationSortir.objects.dates('date_de_sortie', 'year')
    all_years.update(d.year for d in sortie_years)
    
    # Ajouter l'année courante si elle n'est pas déjà présente
    all_years.add(current_year)
    
    return sorted(list(all_years), reverse=True)

@login_required(login_url="accounts:login_user")
def details_entrees(request):
    """Vue détaillée des entrées par mois"""
    # Récupérer l'année sélectionnée ou utiliser l'année courante
    selected_year = int(request.GET.get('year', timezone.now().year))

    # Récupérer le nom de la catégorie depuis la requête GET
    categorie_name = request.GET.get('categorie')
    # Récupérer l'ID de la catégorie correspondante
    categorie_id = Categorie.objects.filter(name=categorie_name, type='entree').values_list('id', flat=True).first() if categorie_name else None

    # Filtrer les entrées par catégorie si l'ID est valide
    if categorie_id:
        entrees = OperationEntrer.objects.filter(
            date_transaction__year=selected_year,
            categorie_id=categorie_id
        ).annotate(
            mois=TruncMonth('date_transaction')
        ).values('mois').annotate(
            total=Sum('montant'),
            nombre_operations=Count('id')
        ).order_by('-mois')
    else:
        # Filtrer les entrées pour l'année sélectionnée
        entrees = OperationEntrer.objects.filter(
            date_transaction__year=selected_year
        ).annotate(
            mois=TruncMonth('date_transaction')
        ).values('mois').annotate(
            total=Sum('montant'),
            nombre_operations=Count('id')
        ).order_by('-mois')

    for entree in entrees:
        entree['operations'] = OperationEntrer.objects.filter(
            date_transaction__month=entree['mois'].month,
            date_transaction__year=entree['mois'].year,
            categorie_id=categorie_id
        ).order_by('-date_transaction') if categorie_id else OperationEntrer.objects.filter(
            date_transaction__month=entree['mois'].month,
            date_transaction__year=entree['mois'].year
        ).order_by('-date_transaction')
        entree['mois_format'] = format_date(entree['mois'], format='MMMM yyyy', locale='fr_FR')

    context = {
        'entrees': entrees,
        'total_general': sum(entree['total'] for entree in entrees),
        'selected_year': selected_year,
        'available_years': get_available_years(),
    }
    return render(request, 'caisse/details/details_entrees.html', context)

@login_required(login_url="accounts:login_user")
def details_sorties(request):
    """Vue détaillée des sorties par mois"""
    selected_year = int(request.GET.get('year', timezone.now().year))

    # Récupérer le nom de la catégorie depuis la requête GET
    categorie_name = request.GET.get('categorie')
    # Récupérer l'ID de la catégorie correspondante
    categorie_id = Categorie.objects.filter(name=categorie_name, type='sortie').values_list('id', flat=True).first() if categorie_name else None

    # Filtrer les sorties par catégorie si l'ID est valide
    if categorie_id:
        sorties = OperationSortir.objects.filter(
            date_de_sortie__year=selected_year,
            categorie_id=categorie_id
        ).annotate(
            mois=TruncMonth('date_de_sortie')
        ).values('mois').annotate(
            total=Sum('montant'),
            nombre_operations=Count('id')
        ).order_by('-mois')
    else:
        sorties = OperationSortir.objects.filter(
            date_de_sortie__year=selected_year
        ).annotate(
            mois=TruncMonth('date_de_sortie')
        ).values('mois').annotate(
            total=Sum('montant'),
            nombre_operations=Count('id')
        ).order_by('-mois')

    for sortie in sorties:
        sortie['operations'] = OperationSortir.objects.filter(
            date_de_sortie__month=sortie['mois'].month,
            date_de_sortie__year=sortie['mois'].year,
            categorie_id=categorie_id
        ).order_by('-date_de_sortie') if categorie_id else OperationSortir.objects.filter(
            date_de_sortie__month=sortie['mois'].month,
            date_de_sortie__year=sortie['mois'].year
        ).order_by('-date_de_sortie')
        sortie['mois_format'] = format_date(sortie['mois'], format='MMMM yyyy', locale='fr_FR')

    context = {
        'sorties': sorties,
        'total_general': sum(sortie['total'] for sortie in sorties),
        'selected_year': selected_year,
        'available_years': get_available_years(),
    }
    return render(request, 'caisse/details/details_sorties.html', context)

@login_required(login_url="accounts:login_user")
def details_solde(request):
    """Vue détaillée du solde par mois"""
    selected_year = int(request.GET.get('year', timezone.now().year))
    
    # Calculer le solde initial (avant l'année sélectionnée)
    solde_initial = (
        OperationEntrer.objects.filter(
            date_transaction__lt=f"{selected_year}-01-01"
        ).aggregate(Sum('montant'))['montant__sum'] or Decimal('0')
    ) - (
        OperationSortir.objects.filter(
            date_de_sortie__lt=f"{selected_year}-01-01"
        ).aggregate(Sum('montant'))['montant__sum'] or Decimal('0')
    )

    # Calculer les entrées et sorties pour l'année sélectionnée
    entrees = OperationEntrer.objects.filter(
        date_transaction__year=selected_year
    ).annotate(
        mois=TruncMonth('date_transaction')
    ).values('mois').annotate(
        total_entrees=Sum('montant')
    ).order_by('mois')  # Changer en ordre croissant

    sorties = OperationSortir.objects.filter(
        date_de_sortie__year=selected_year
    ).annotate(
        mois=TruncMonth('date_de_sortie')
    ).values('mois').annotate(
        total_sorties=Sum('montant')
    ).order_by('mois')  # Changer en ordre croissant

    # Créer des dictionnaires pour un accès facile
    entrees_dict = {e['mois']: e['total_entrees'] or Decimal('0') for e in entrees}
    sorties_dict = {s['mois']: s['total_sorties'] or Decimal('0') for s in sorties}
    
    # Obtenir tous les mois uniques et les trier par ordre croissant
    tous_mois = sorted(set(list(entrees_dict.keys()) + list(sorties_dict.keys())))
    
    # Calculer les soldes cumulatifs
    soldes_mensuels = []
    solde_cumule = solde_initial
    total_entrees_annee = Decimal('0')
    total_sorties_annee = Decimal('0')
    
    # Premier passage : calculer les soldes cumulés dans l'ordre chronologique
    soldes_temp = {}
    for mois in tous_mois:
        entrees_mois = entrees_dict.get(mois, Decimal('0'))
        sorties_mois = sorties_dict.get(mois, Decimal('0'))
        solde_mois = entrees_mois - sorties_mois
        solde_cumule += solde_mois
        
        total_entrees_annee += entrees_mois
        total_sorties_annee += sorties_mois
        
        soldes_temp[mois] = {
            'mois': mois,
            'mois_format': format_date(mois, format='MMMM yyyy', locale='fr_FR'),
            'entrees': entrees_mois,
            'sorties': sorties_mois,
            'solde_mois': solde_mois,
            'solde_cumule': solde_cumule
        }
    
    # Deuxième passage : créer la liste finale dans l'ordre décroissant
    for mois in reversed(tous_mois):
        soldes_mensuels.append(soldes_temp[mois])

    context = {
        'soldes': soldes_mensuels,
        'total_entrees': total_entrees_annee,
        'total_sorties': total_sorties_annee,
        'solde_final': solde_cumule,
        'selected_year': selected_year,
        'available_years': get_available_years(),
    }
    return render(request, 'caisse/details/details_solde.html', context)

@login_required(login_url="accounts:login_user")
def ajouter_element(request):
    element_type = request.GET.get('type')
    return_url = request.GET.get('return_url')
    
    if element_type == 'catégorie':
        if request.method == 'POST':
            form = CategorieForm(request.POST)
            if form.is_valid():
                categorie = form.save()
                UserActivity.objects.create(
                    user=request.user, 
                    action='Création', 
                    description='a créé une nouvelle catégorie'
                )
                messages.success(request, "Catégorie ajoutée avec succès")
                if return_url:
                    return redirect(return_url)
                return redirect('caisse:liste_categories')
        else:
            form = CategorieForm()
        return render(request, 'caisse/parametres/ajouter_categorie.html', {'form': form, 'return_url': return_url})
    
    elif element_type == 'bénéficiaire':
        if request.method == 'POST':
            form = BeneficiaireForm(request.POST)
            if form.is_valid():
                beneficiaire = form.save()
                UserActivity.objects.create(
                    user=request.user, 
                    action='Création', 
                    description='a créé un nouveau bénéficiaire'
                )
                messages.success(request, "Bénéficiaire ajouté avec succès")
                if return_url:
                    return redirect(return_url)
                return redirect('caisse:liste_beneficiaires')
        else:
            form = BeneficiaireForm()
        return render(request, 'caisse/acteurs/ajouter_beneficiaire.html', {'form': form, 'return_url': return_url})
    
    elif element_type == 'fournisseur':
        if request.method == 'POST':
            form = FournisseurForm(request.POST)
            if form.is_valid():
                fournisseur = form.save()
                UserActivity.objects.create(
                    user=request.user, 
                    action='Création', 
                    description='a créé un nouveau fournisseur'
                )
                messages.success(request, "Fournisseur ajouté avec succès")
                if return_url:
                    return redirect(return_url)
                return redirect('caisse:liste_fournisseurs')
        else:
            form = FournisseurForm()
        return render(request, 'caisse/acteurs/ajouter_fournisseur.html', {'form': form, 'return_url': return_url})
    
    return redirect('caisse:index')

@login_required(login_url="accounts:login_user")
def verifier_categorie(request, id):
    try:
        exists = Categorie.objects.filter(id=id).exists()
        return JsonResponse({'exists': exists})
    except:
        return JsonResponse({'exists': False})

@login_required(login_url="accounts:login_user")
def verifier_beneficiaire(request, id):
    try:
        exists = Beneficiaire.objects.filter(id=id).exists()
        return JsonResponse({'exists': exists})
    except:
        return JsonResponse({'exists': False})

@login_required(login_url="accounts:login_user")
def verifier_fournisseur(request, id):
    try:
        exists = Fournisseur.objects.filter(id=id).exists()
        return JsonResponse({'exists': exists})
    except:
        return JsonResponse({'exists': False})